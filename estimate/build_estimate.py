# -*- coding: utf-8 -*-
"""聖建様向け 特別教育eラーニング 概算見積シート"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

F = "Arial"
NAVY = "16324F"; AMBER = "E8960C"; LIGHT = "F3F5F7"
YEL = PatternFill("solid", fgColor="FFF2CC")
HDR = PatternFill("solid", fgColor=NAVY)
SUB = PatternFill("solid", fgColor="DCE6F1")
GREY = PatternFill("solid", fgColor=LIGHT)
thin = Side(style="thin", color="BFC9D1")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
YEN = '¥#,##0;(¥#,##0);-'
HRS = '0.0"時間";;-'

def hdr(ws, row, vals, widths=None, height=22):
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=F, bold=True, color="FFFFFF", size=10)
        c.fill = HDR; c.border = BOX
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = height
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

def title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = Font(name=F, bold=True, size=16, color=NAVY)
    ws.row_dimensions[1].height = 26
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(name=F, size=9.5, color="6B7B8A")

def cell(ws, r, c, v, *, bold=False, fmt=None, fill=None, align=None,
         color="000000", size=10, wrap=False, border=True):
    x = ws.cell(row=r, column=c, value=v)
    x.font = Font(name=F, bold=bold, size=size, color=color)
    if fmt: x.number_format = fmt
    if fill: x.fill = fill
    if border: x.border = BOX
    x.alignment = Alignment(horizontal=align or "left", vertical="center", wrap_text=wrap)
    return x

wb = openpyxl.Workbook()

# =====================================================================
# 02_科目マスタ  (built first: other sheets reference it)
# =====================================================================
ws = wb.active; ws.title = "02_科目マスタ"
title(ws, "特別教育 科目別 法定時間マスタ",
      "出典：安全衛生特別教育規程（昭和47年労働省告示第92号）ほか。学科時間が、eラーニング動画の尺を設計する基準になります。")

SUBJECTS = [
    # 業務名, 学科h, 実技h, eラーニング適性, 確認状況・備考
    ("足場の組立て等の業務", 6.0, 0.0, "◎ 学科のみで完結",
     "実技の定めなし。オンラインで修了まで完結できる有力候補"),
    ("粉じん作業に係る業務", 4.5, 0.0, "◎ 学科のみで完結",
     "じん肺法・粉じん障害防止規則関連。実技の定めなし"),
    ("石綿等が使用されている建築物等の解体等の業務", 4.5, 0.0, "◎ 学科のみで完結",
     "石綿障害予防規則。解体・改修需要が継続的にある"),
    ("酸素欠乏・硫化水素危険作業（第2種）", 5.5, 0.0, "◎ 学科のみで完結",
     "酸素欠乏危険作業特別教育規程。第1種は科目構成が異なるため要原典確認"),
    ("フルハーネス型墜落制止用器具を用いて行う作業", 4.5, 1.5, "○ 学科のみ提供",
     "実技1.5時間は受講者の所属事業者が実施。Web講座の標準的な提供形態"),
    ("テールゲートリフターの操作の業務", 4.0, 2.0, "○ 学科のみ提供",
     "2024年2月1日より義務化。新規需要が最も大きい科目"),
    ("自由研削といしの取替え等の業務", 4.0, 3.0, "○ 学科のみ提供",
     "製造業・建設業ともに対象者が多い"),
    ("低圧電気取扱業務（開閉器の操作のみ）", 7.0, 1.0, "○ 学科のみ提供",
     "実技は1時間以上。学科が7時間と長く、制作費は上振れする"),
    ("低圧電気取扱業務（充電電路の敷設等）", 7.0, 7.0, "△ 学科のみ提供",
     "実技7時間以上。実技負担が大きく受講者側のハードルが高い"),
    ("高所作業車の運転（作業床の高さ10m未満）", 6.0, 3.0, "○ 学科のみ提供",
     "実技3時間は実機が必要"),
    ("玉掛けの業務（つり上げ荷重1トン未満）", 5.0, 4.0, "○ 学科のみ提供",
     "実技4時間は実機が必要"),
    ("小型車両系建設機械の運転（機体重量3トン未満）", 7.0, 6.0, "△ 学科のみ提供",
     "実技6時間。整地・運搬・積込み用および掘削用"),
    ("チェーンソーを用いて行う伐木等の業務", 9.0, 8.0, "△ 学科のみ提供",
     "令和元年改正で統合・拡充。建設業では対象者が限定的"),
    ("アーク溶接等の業務", 11.0, 10.0, "△ 学科のみ提供",
     "学科11時間・実技10時間と最大級。制作費が最も高くなる"),
]
REF = [
    ("【参考】職長・安全衛生責任者教育", 14.0, 0.0, "参考（特別教育ではない）",
     "労働安全衛生法第60条。安衛則40条の12時間以上に、安全衛生責任者分を加えた運用が一般的。要原典確認"),
    ("【参考】有機溶剤業務従事者 労働衛生教育", 4.5, 0.0, "参考（特別教育ではない）",
     "通達に基づく安全衛生教育。特別教育ではないため罰則の対象外。時間は実施機関により差がある"),
]

hdr(ws, 4, ["業務の名称", "学科\n(時間)", "実技\n(時間)", "合計\n(時間)",
            "eラーニング適性", "確認状況・備考"],
    widths=[44, 8, 8, 8, 20, 56], height=32)

r = 5
for name, gh, jh, fit, note in SUBJECTS:
    cell(ws, r, 1, name, wrap=True)
    cell(ws, r, 2, gh, fmt=HRS, align="center", bold=True, color="B87400")
    cell(ws, r, 3, jh, fmt=HRS, align="center")
    cell(ws, r, 4, f"=B{r}+C{r}", fmt=HRS, align="center")
    cell(ws, r, 5, fit, align="center", size=9.5)
    cell(ws, r, 6, note, wrap=True, size=9, color="4A5A68")
    ws.row_dimensions[r].height = 28
    r += 1

r += 1
cell(ws, r, 1, "参考：特別教育以外の教育（同じ仕組みで配信できます）", bold=True,
     color=NAVY, fill=SUB, border=False)
r += 1
for name, gh, jh, fit, note in REF:
    cell(ws, r, 1, name, wrap=True, color="4A5A68")
    cell(ws, r, 2, gh, fmt=HRS, align="center", color="4A5A68")
    cell(ws, r, 3, jh, fmt=HRS, align="center", color="4A5A68")
    cell(ws, r, 4, f"=B{r}+C{r}", fmt=HRS, align="center", color="4A5A68")
    cell(ws, r, 5, fit, align="center", size=9.5, color="4A5A68")
    cell(ws, r, 6, note, wrap=True, size=9, color="4A5A68")
    ws.row_dimensions[r].height = 28
    r += 1

r += 2
notes = [
    "■ この表の読み方",
    "・「学科」＝eラーニングで提供できる時間。この時間数が、制作する動画の尺を決める基準になります。",
    "・「実技」＝オンラインで代替できません。受講者の所属事業者が自ら実施する前提です（フルハーネス等のWeb講座で一般的な提供形態）。",
    "・法定時間は「以上」の定めです。動画の総再生時間だけでなく、章末テスト・演習時間を含めた総学習時間で満たす設計にします。",
    "",
    "■ 出典と確認状況",
    "・時間数は安全衛生特別教育規程（昭和47年労働省告示第92号）および各講習機関の公表カリキュラムに基づきます。",
    "・複数の講習機関の公表値が一致することを確認していますが、契約前に規程の原典で科目ごとに確定させてください。",
    "・「要原典確認」と記載した科目は、区分により時間が異なるため特に注意が必要です。",
]
for t in notes:
    c = cell(ws, r, 1, t, border=False, size=9.5,
             bold=t.startswith("■"), color=NAVY if t.startswith("■") else "4A5A68")
    r += 1
ws.freeze_panes = "A5"

# =====================================================================
# 03_教材制作費
# =====================================================================
ws = wb.create_sheet("03_教材制作費")
title(ws, "教材制作費の試算",
      "黄色のセルが入力欄です。A列の「制作対象」に 1 を入れた科目だけが集計されます。")

cell(ws, 4, 1, "単価設定（黄色セルを編集してください）", bold=True, color=NAVY, fill=SUB, border=False)
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 44
for col, w in zip("CDEFG", [11, 11, 15, 15, 15]):
    ws.column_dimensions[col].width = w
ws.column_dimensions["H"].width = 16

unit_rows = [
    ("教材制作単価（学科1時間あたり）", 250000,
     "構成台本の書き起こし、スライド制作、図解アニメーション、AI音声ナレーション、バーチャル講師映像、編集、書き出しまで。撮影は行いません。"),
    ("監修費（1科目あたり）", 0,
     "聖建様の社内有資格者が監修される前提のため 0 円としています。外部の労働安全衛生コンサルタント等に依頼される場合は 350,000 円程度を目安に入力してください。監修者の氏名・保有資格はシート06の台帳に記録します。"),
    ("理解度テスト制作（1科目あたり）", 120000,
     "章末テストおよび修了確認テストの問題作成・解説文の作成。"),
]
r = 5
for label, val, note in unit_rows:
    cell(ws, r, 2, label, bold=True)
    c = cell(ws, r, 3, val, fmt=YEN, fill=YEL, align="right", bold=True, color="0000FF")
    c.comment = Comment(note, "見積前提")
    cell(ws, r, 4, "", border=True)
    cell(ws, r, 5, note, wrap=True, size=8.5, color="4A5A68")
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 34
    r += 1

cell(ws, 9, 2, "参考：制作グレード別の単価目安（学科1時間あたり）", bold=True, color=NAVY, border=False)
grades = [
    ("ライト", 150000, "スライド＋AI音声ナレーションのみ。図解は静止画。最も安価だが受講者の離脱率が上がりやすい。"),
    ("スタンダード（推奨）", 250000, "スライド＋図解アニメーション＋AI音声＋バーチャル講師。長時間の視聴に耐える構成。"),
    ("プレミアム", 400000, "実写素材・3D図解・章ごとの演習を含む。差別化を最優先する場合。"),
]
hdr(ws, 10, ["", "グレード", "単価/時間", "", "内容"], height=20)
r = 11
for g, v, note in grades:
    cell(ws, r, 2, g, bold=(g.startswith("スタンダード")))
    cell(ws, r, 3, v, fmt=YEN, align="right")
    cell(ws, r, 4, "")
    cell(ws, r, 5, note, wrap=True, size=8.5, color="4A5A68")
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 26
    r += 1

TOP = 16
cell(ws, TOP - 1, 1, "科目別の試算（A列に 1 を入力した科目のみ集計されます）",
     bold=True, color=NAVY, fill=SUB, border=False)
hdr(ws, TOP, ["制作対象\n(1で選択)", "業務の名称", "学科\n(時間)", "実技\n(時間)",
              "教材制作費", "監修費", "テスト制作費", "科目 小計"], height=32)

first = TOP + 1
n = len(SUBJECTS)
for i, (name, gh, jh, fit, note) in enumerate(SUBJECTS):
    r = first + i
    src = 5 + i  # matching row on 02_科目マスタ
    flag = 1 if name.startswith(("足場", "フルハーネス", "テールゲート")) else 0
    cell(ws, r, 1, flag, fill=YEL, align="center", bold=True, color="0000FF")
    cell(ws, r, 2, f"='02_科目マスタ'!A{src}", wrap=True, color="008000")
    cell(ws, r, 3, f"='02_科目マスタ'!B{src}", fmt=HRS, align="center", color="008000")
    cell(ws, r, 4, f"='02_科目マスタ'!C{src}", fmt=HRS, align="center", color="008000")
    cell(ws, r, 5, f"=IF($A{r}=1,$C{r}*$C$5,0)", fmt=YEN, align="right")
    cell(ws, r, 6, f"=IF($A{r}=1,$C$6,0)", fmt=YEN, align="right")
    cell(ws, r, 7, f"=IF($A{r}=1,$C$7,0)", fmt=YEN, align="right")
    cell(ws, r, 8, f"=SUM(E{r}:G{r})", fmt=YEN, align="right", bold=True)
    ws.row_dimensions[r].height = 26

last = first + n - 1
tr = last + 1
cell(ws, tr, 1, f"=SUM(A{first}:A{last})", align="center", bold=True, fill=SUB)
cell(ws, tr, 2, "合計（選択した科目数・時間・費用）", bold=True, fill=SUB)
cell(ws, tr, 3, f"=SUMPRODUCT($A{first}:$A{last},$C{first}:$C{last})", fmt=HRS, align="center", bold=True, fill=SUB)
cell(ws, tr, 4, "", fill=SUB)
for col in (5, 6, 7, 8):
    L = get_column_letter(col)
    cell(ws, tr, col, f"=SUM({L}{first}:{L}{last})", fmt=YEN, align="right", bold=True, fill=SUB)
ws.row_dimensions[tr].height = 24

r = tr + 2
for t in [
    "■ 教材制作の進め方（撮影は行いません）",
    "1. 法定の科目・範囲・時間を規程の原典から書き出し、章立てと各章の尺を設計します（ここで法定時間の充足を担保します）。",
    "2. 構成台本を一から書き起こします。他社の講義映像や市販冊子をそのまま使うことは著作権法上できません。",
    "3. スライド・図解を制作し、AI音声ナレーションとバーチャル講師映像を合成します。",
    "4. 章末テストと修了確認テストを作成します。",
    "5. 聖建様の社内有資格者が監修し、内容が現場の実態と合っているかを確認します（法令への当てはめは弊社が行います）。",
    "6. 監修者の氏名・保有資格・レビュー日をシート06の台帳に記録し、監督署への説明資料とします。",
    "",
    "■ 2科目目以降について",
    "・章立てのテンプレート、スライドのデザイン、動画の書き出し設定は1科目目で確立するため、2科目目以降は制作単価を抑えられます。",
    "・目安として、2科目目以降は上表の8割程度を想定してください。",
    "",
    "■ 監修について（2026年8月30日の打合せを反映）",
    "・聖建様の社内に、科目に対応する有資格者・実務経験者がいらっしゃるため、外部コンサルタントの監修費は計上していません。",
    "・かわりに、(1) 監修者の資格を修了証の写しで確認、(2) 章立て設計の段階で所轄労働基準監督署へ事前相談（無料）、の2点を実施します。",
]:
    cell(ws, r, 1, t, border=False, size=9.5,
         bold=t.startswith("■"), color=NAVY if t.startswith("■") else "4A5A68")
    r += 1
ws.freeze_panes = f"A{TOP+1}"

# =====================================================================
# 04_システム費用
# =====================================================================
ws = wb.create_sheet("04_システム費用")
title(ws, "配信システムの概算費用",
      "実技会場の予約機能が不要になったため、前回ご提示分より構成を絞っています。")
for col, w in zip("ABCDE", [10, 34, 16, 16, 60]):
    ws.column_dimensions[col].width = w

hdr(ws, 4, ["区分", "機能", "概算(下限)", "概算(上限)", "内容"], height=22)

SYS = [
    ("フェーズ1", "要件定義・基本設計", 400000, 700000,
     "業務フローの確定、画面設計、利用規約・特定商取引法表記・個人情報の取扱い方針の整備支援"),
    ("フェーズ1", "受講者ポータル", 500000, 800000,
     "アカウント登録、マイページ、受講状況の表示、スマートフォン対応"),
    ("フェーズ1", "本人確認・顔照合", 450000, 800000,
     "本人確認書類のアップロード、顔写真の登録、受講中のランダム自撮り撮影と照合、運営による承認フロー"),
    ("フェーズ1", "動画配信・視聴ログ", 400000, 700000,
     "学科動画の配信、スキップ・早送りの禁止、視聴時間の秒単位記録（法定時間の充足を証明）"),
    ("フェーズ1", "テスト・自動採点", 250000, 450000,
     "章末テスト（不合格なら次章に進めない制御）、修了確認テスト、自動採点"),
    ("フェーズ1", "修了証の発行", 200000, 350000,
     "PDF修了証の自動発行、再ダウンロード、発行履歴の管理"),
    ("フェーズ2", "企業管理ポータル", 400000, 700000,
     "発注企業のご担当者向け。従業員の一括申込、受講進捗の確認、修了者一覧のCSV出力"),
    ("フェーズ2", "決済（カード・請求書）", 250000, 450000,
     "クレジットカード決済、請求書払い、領収書の発行"),
    ("フェーズ2", "運営管理画面の拡張", 350000, 600000,
     "本人確認の承認キュー、顔照合アラート、受講ログの照会、売上・入金の管理"),
    ("フェーズ2", "記録保存・CSV出力", 200000, 350000,
     "労働安全衛生規則第38条に対応した3年以上の自動保存、一括出力"),
]
r = 5
p1 = []; p2 = []
for ph, fn, lo, hi, note in SYS:
    cell(ws, r, 1, ph, align="center", bold=True,
         color=NAVY if ph == "フェーズ1" else "6B7B8A")
    cell(ws, r, 2, fn, wrap=True)
    cell(ws, r, 3, lo, fmt=YEN, align="right", fill=YEL, color="0000FF")
    cell(ws, r, 4, hi, fmt=YEN, align="right", fill=YEL, color="0000FF")
    cell(ws, r, 5, note, wrap=True, size=9, color="4A5A68")
    ws.row_dimensions[r].height = 30
    (p1 if ph == "フェーズ1" else p2).append(r)
    r += 1

def subtotal(row, label, rows, fill):
    cell(ws, row, 1, "", fill=fill)
    cell(ws, row, 2, label, bold=True, fill=fill)
    cell(ws, row, 3, f"=SUM(C{rows[0]}:C{rows[-1]})", fmt=YEN, align="right", bold=True, fill=fill)
    cell(ws, row, 4, f"=SUM(D{rows[0]}:D{rows[-1]})", fmt=YEN, align="right", bold=True, fill=fill)
    cell(ws, row, 5, "", fill=fill)

r1 = r
subtotal(r1, "フェーズ1 小計（まずここから始めます）", p1, SUB)
r2 = r1 + 1
subtotal(r2, "フェーズ2 小計（運用開始後に追加）", p2, SUB)
r3 = r2 + 1
cell(ws, r3, 1, "", fill=GREY)
cell(ws, r3, 2, "システム初期費用 合計", bold=True, fill=GREY)
cell(ws, r3, 3, f"=C{r1}+C{r2}", fmt=YEN, align="right", bold=True, fill=GREY)
cell(ws, r3, 4, f"=D{r1}+D{r2}", fmt=YEN, align="right", bold=True, fill=GREY)
cell(ws, r3, 5, "", fill=GREY)

mr = r3 + 3
cell(ws, mr - 1, 1, "月額運用費", bold=True, color=NAVY, fill=SUB, border=False)
hdr(ws, mr, ["", "項目", "月額(下限)", "月額(上限)", "内容"], height=20)
MONTH = [
    ("サーバー・インフラ", 20000, 40000, "クラウド利用料。受講者数の増加に応じて変動します"),
    ("保守・障害対応", 30000, 50000, "監視、障害対応、軽微な改修"),
    ("本人確認API（従量）", 0, 0, "1件あたり100〜200円程度。受講申込者数に比例します"),
    ("動画配信（従量）", 0, 0, "視聴時間に比例。学科が長い科目ほど増えます"),
]
r = mr + 1
mrows = []
for it, lo, hi, note in MONTH:
    cell(ws, r, 1, "")
    cell(ws, r, 2, it)
    cell(ws, r, 3, lo, fmt=YEN, align="right", fill=YEL, color="0000FF")
    cell(ws, r, 4, hi, fmt=YEN, align="right", fill=YEL, color="0000FF")
    cell(ws, r, 5, note, wrap=True, size=9, color="4A5A68")
    mrows.append(r); r += 1
cell(ws, r, 1, "", fill=SUB)
cell(ws, r, 2, "月額 固定費 合計（従量分を除く）", bold=True, fill=SUB)
cell(ws, r, 3, f"=SUM(C{mrows[0]}:C{mrows[-1]})", fmt=YEN, align="right", bold=True, fill=SUB)
cell(ws, r, 4, f"=SUM(D{mrows[0]}:D{mrows[-1]})", fmt=YEN, align="right", bold=True, fill=SUB)
cell(ws, r, 5, "", fill=SUB)
SYS_LO, SYS_HI, MON_LO, MON_HI = f"C{r3}", f"D{r3}", f"C{r}", f"D{r}"

r += 2
for t in [
    "■ 前回ご提示分からの変更点",
    "・実技会場の予約機能、出欠管理機能を削除しました（実技は受講者の所属事業者が実施するため）。",
    "・学科のみの提供に絞ったことで、フェーズ1の範囲が明確になり、着手しやすい構成になっています。",
]:
    cell(ws, r, 1, t, border=False, size=9.5,
         bold=t.startswith("■"), color=NAVY if t.startswith("■") else "4A5A68")
    r += 1

# =====================================================================
# 01_概算サマリー
# =====================================================================
ws = wb.create_sheet("01_概算サマリー", 0)
title(ws, "特別教育 eラーニング事業  全体概算",
      "株式会社聖建 御中 ／ 株式会社スペリオル ／ 2026年8月30日時点の概算")
for col, w in zip("ABCDE", [6, 40, 18, 18, 58]):
    ws.column_dimensions[col].width = w

hdr(ws, 4, ["", "項目", "概算(下限)", "概算(上限)", "備考"], height=22)
MT = "'03_教材制作費'"
ST = "'04_システム費用'"
rows = [
    ("教材制作費（選択した科目の合計）", f"={MT}!H{tr}", f"={MT}!H{tr}",
     "シート03で科目を選択すると自動で反映されます。上限・下限は同額（単価は1本）"),
    ("システム初期費用 フェーズ1", f"={ST}!C{r1}", f"={ST}!D{r1}",
     "受講者ポータル、本人確認、動画配信、テスト、修了証。まずここから"),
    ("システム初期費用 フェーズ2", f"={ST}!C{r2}", f"={ST}!D{r2}",
     "企業管理ポータル、決済、運営管理、記録保存。運用開始後に追加"),
]
r = 5
srows = []
for label, lo, hi, note in rows:
    cell(ws, r, 1, "")
    cell(ws, r, 2, label, bold=True)
    cell(ws, r, 3, lo, fmt=YEN, align="right", color="008000")
    cell(ws, r, 4, hi, fmt=YEN, align="right", color="008000")
    cell(ws, r, 5, note, wrap=True, size=9, color="4A5A68")
    ws.row_dimensions[r].height = 30
    srows.append(r); r += 1

cell(ws, r, 1, "", fill=GREY)
cell(ws, r, 2, "初期投資 合計", bold=True, size=11.5, fill=GREY)
cell(ws, r, 3, f"=SUM(C{srows[0]}:C{srows[-1]})", fmt=YEN, align="right", bold=True, size=11.5, fill=GREY)
cell(ws, r, 4, f"=SUM(D{srows[0]}:D{srows[-1]})", fmt=YEN, align="right", bold=True, size=11.5, fill=GREY)
cell(ws, r, 5, "", fill=GREY)
ws.row_dimensions[r].height = 26
TOTROW = r

r += 1
cell(ws, r, 1, "")
cell(ws, r, 2, "うち フェーズ1 までで着手する場合", bold=True)
cell(ws, r, 3, f"=C{srows[0]}+C{srows[1]}", fmt=YEN, align="right", bold=True, color="B87400")
cell(ws, r, 4, f"=D{srows[0]}+D{srows[1]}", fmt=YEN, align="right", bold=True, color="B87400")
cell(ws, r, 5, "教材とフェーズ1のみ。まずこの範囲でのご検討を推奨します", wrap=True, size=9, color="4A5A68")
ws.row_dimensions[r].height = 26

r += 2
cell(ws, r, 2, "月額運用費（固定分）", bold=True)
cell(ws, r, 3, f"={ST}!{MON_LO}", fmt=YEN, align="right", color="008000")
cell(ws, r, 4, f"={ST}!{MON_HI}", fmt=YEN, align="right", color="008000")
cell(ws, r, 5, "別途、本人確認APIと動画配信の従量課金が発生します", wrap=True, size=9, color="4A5A68")

r += 3
cell(ws, r, 2, "選択中の科目数", bold=True)
cell(ws, r, 3, f"={MT}!A{tr}", align="center", bold=True, color="008000")
cell(ws, r, 4, "科目", align="left", size=9)
cell(ws, r, 5, "シート03のA列で選択します", size=9, color="4A5A68")
r += 1
cell(ws, r, 2, "選択中の学科時間 合計", bold=True)
cell(ws, r, 3, f"={MT}!C{tr}", fmt=HRS, align="center", bold=True, color="008000")
cell(ws, r, 4, "", align="left")
cell(ws, r, 5, "制作する動画の総尺の設計基準になります", size=9, color="4A5A68")

r += 3
for t in [
    "■ このシートの使い方",
    "・シート03「教材制作費」のA列に 1 を入れると、その科目が集計に入ります。初期状態では足場・フルハーネス・テールゲートリフターの3科目を選択しています。",
    "・単価（黄色のセル）を変更すると、このサマリーまで自動で再計算されます。",
    "・緑色の数字は他シートからの参照、青色の数字は入力値、黒色は計算結果です。",
    "",
    "■ 重要：この金額は概算です",
    "・単価は同種案件の一般的な相場に基づく想定値です。正式なお見積りは要件確定後に別途ご提示します。",
    "・シート05に、金額の前提と責任の分担範囲を記載しています。必ずあわせてご確認ください。",
]:
    cell(ws, r, 2, t, border=False, size=9.5,
         bold=t.startswith("■"), color=NAVY if t.startswith("■") else "4A5A68")
    r += 1

# =====================================================================
# 05_前提・責任分界
# =====================================================================
ws = wb.create_sheet("05_前提・責任分界")
title(ws, "前提条件と、責任の分担範囲",
      "ご契約前に、必ずこの内容についてご合意ください。")
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 120

BLOCKS = [
    ("■ 1. 保証する範囲と、保証しない範囲", NAVY, [
        ("○ 弊社が保証すること", "008000"),
        ("　・教材が、安全衛生特別教育規程が定める科目・範囲・時間を満たしていること。", None),
        ("　・その確認を、労働安全衛生の専門家（監修者）が、氏名と資格を明示したうえで行うこと。", None),
        ("　・受講の記録（受講日時、視聴時間、本人確認の結果、テストの得点）を、証明可能な形で保存・出力できること。", None),
        ("", None),
        ("× 弊社が保証できないこと", "B3402F"),
        ("　・受講者の所属事業者が負う、労働安全衛生法第59条第3項の特別教育の実施義務が果たされたこと。", None),
        ("　　この義務は法律上、受講者を雇用する事業者に課されるものであり、教材の提供者が代わりに負うことはできません。", None),
        ("　・受講者が業務に従事した際に労働災害が発生しないこと。", None),
        ("　・監督署その他の行政機関が、個別の事案において本教材の内容を適法と判断すること。", None),
    ]),
    ("■ 2. 教材内容の監修について（社内監修を前提とします）", NAVY, [
        ("・聖建様の社内に、科目に対応する有資格者・実務経験者がいらっしゃるため、外部コンサルタントの監修は不要と判断しました。", None),
        ("・弊社が担うのは「法令への当てはめ」です。安全衛生特別教育規程の原典から科目・範囲・時間を書き出し、それを満たす章立てと尺を設計します。", None),
        ("・聖建様の監修者に担っていただくのは「内容が現場の実態と合っているか」の確認です。法令の条文を調べていただく必要はありません。", None),
        ("・監修者の氏名・保有資格・修了証の確認状況・台本のレビュー日を、シート06の台帳に記録します。これが監督署への説明資料になります。", None),
        ("", None),
        ("※ 注意：特別教育を「受講した経験」だけでは、講師要件の裏付けとしては弱くなります。", "B3402F"),
        ("　 科目に対応する作業主任者技能講習の修了証、または免許の保有を、現物の写しでご確認ください（費用はかかりません）。", "B3402F"),
        ("", None),
        ("・弁護士による確認では代替できません。労働安全衛生教育の内容の妥当性は、法律実務ではなく安全衛生の専門領域だからです。", None),
        ("・章立ての設計が固まった段階で、所轄の労働基準監督署（半田市の事業場は半田労働基準監督署）へ事前相談を行います。無料です。", "008000"),
        ("　 特別教育には登録・認可の制度がないため、事前にお墨付きを得る窓口は存在しません。相談の記録を残すことが、実務上の最善策になります。", "008000"),
    ]),
    ("■ 3. 著作権に関する前提", NAVY, [
        ("・他社が実施する講習の映像を撮影して教材化することは、著作権法上できません。受講規約への違反にもあたります。", "B3402F"),
        ("・市販の冊子・テキストを原稿として利用する場合は、発行元の許諾が必要です。許諾なしでの利用はできません。", "B3402F"),
        ("・本お見積りは、構成台本を一から書き起こす前提です。そのため、完成した教材は聖建様が自由に利用・改訂できる資産になります。", "008000"),
    ]),
    ("■ 4. 法定教育時間の充足について", NAVY, [
        ("・法定時間は「以上」の定めです。動画の総再生時間だけで満たすのではなく、章末テスト・演習を含む総学習時間で設計します。", None),
        ("・章立てと各章の尺は、制作着手前に設計書として確定させ、聖建様のご承認をいただきます。", None),
        ("・この設計を飛ばして制作を始めると、完成後に時間不足が判明し、作り直しが発生します。最も費用が膨らむ失敗パターンです。", "B3402F"),
    ]),
    ("■ 5. 金額の前提", NAVY, [
        ("・本シートの金額は、同種案件の一般的な相場に基づく概算です。正式なお見積りは要件確定後に別途ご提示します。", None),
        ("・教材制作費は「学科1時間あたりの単価 × 法定学科時間」で算出しています。科目の学科時間が長いほど金額は上がります。", None),
        ("・撮影・スタジオ費用は含みません（バーチャル講師およびAI音声ナレーションを用いるため）。", None),
        ("・2科目目以降は、テンプレートの再利用により制作単価を抑えられます（監修費は科目ごとに必要です）。", None),
        ("・受講者への販売価格、収支計画は本シートには含みません。別途ご相談ください。", None),
    ]),
    ("■ 6. 出典", NAVY, [
        ("・安全衛生特別教育規程（昭和47年9月30日 労働省告示第92号）", None),
        ("・酸素欠乏危険作業特別教育規程（昭和47年9月30日 労働省告示第132号）", None),
        ("・労働安全衛生規則 第36条（特別教育を必要とする業務）、第38条（記録の保存・3年間）", None),
        ("・インターネット等を介したeラーニング等による安全衛生教育の実施について（令和3年1月25日 基安安発0125第2号 ほか）", None),
        ("・各科目の時間数は、上記規程および複数の講習機関が公表するカリキュラムで相互に確認しています。", None),
    ]),
]
r = 4
for head, hcol, lines in BLOCKS:
    cell(ws, r, 2, head, bold=True, size=11.5, color=hcol, fill=SUB, border=False)
    ws.row_dimensions[r].height = 22
    r += 1
    for txt, col in lines:
        cell(ws, r, 2, txt, size=9.5, border=False,
             color=col or "1C2A38", bold=bool(col))
        r += 1
    r += 1


# =====================================================================
# 06_講師・監修者台帳
# =====================================================================
ws = wb.create_sheet("06_講師・監修者台帳")
title(ws, "講師・監修者 台帳",
      "特別教育の講師に法定の資格要件はありませんが、「教習科目について十分な知識と経験を有する者」であることの説明が求められます。その説明資料がこの台帳です。")

widths = [26, 14, 16, 24, 16, 24, 10, 12, 12, 14]
hdr(ws, 4, ["科目（業務の名称）", "監修者\n(講師)", "所属・役職",
            "保有資格 ①（科目に対応するもの）", "資格番号・取得年月",
            "保有資格 ②", "実務\n経験年数", "修了証写し\nの確認", "台本レビュー\n完了日", "監修者の署名"],
    widths=widths, height=36)

cell(ws, 5, 1, "【記入例】足場の組立て等の業務", size=9, color="6B7B8A")
example = ["山田 太郎", "聖建 工事部 部長", "足場の組立て等作業主任者技能講習 修了",
           "第12345号 / 2015年6月", "職長・安全衛生責任者教育 修了", 22, "済 (2026-09-10)",
           "2026-10-15", ""]
for i, v in enumerate(example, start=2):
    cell(ws, 5, i, v, size=9, color="6B7B8A", wrap=True,
         align="center" if i in (7, 8, 9, 10) else "left")
ws.row_dimensions[5].height = 30

for i, (name, gh, jh, fit, note) in enumerate(SUBJECTS):
    r = 6 + i
    cell(ws, r, 1, f"='02_科目マスタ'!A{5+i}", wrap=True, color="008000", size=9.5)
    for c in range(2, 11):
        cell(ws, r, c, "", fill=YEL,
             align="center" if c in (7, 8, 9, 10) else "left")
    ws.row_dimensions[r].height = 30

r = 6 + len(SUBJECTS) + 2
GUIDE = [
    ("■ 科目ごとに、どの資格が説明力を持つか", NAVY),
    ("　足場の組立て等　　　　　→　足場の組立て等作業主任者技能講習 修了者", None),
    ("　石綿等の解体等　　　　　→　石綿作業主任者技能講習 修了者", None),
    ("　酸素欠乏・硫化水素　　　→　酸素欠乏・硫化水素危険作業主任者技能講習 修了者", None),
    ("　低圧電気取扱　　　　　　→　第一種／第二種電気工事士、電気主任技術者", None),
    ("　粉じん作業　　　　　　　→　特定粉じん作業に係る実務経験、じん肺関係の教育歴", None),
    ("　フルハーネス　　　　　　→　建築物等の鉄骨の組立て等作業主任者、または高所作業の実務経験＋教育歴", None),
    ("　テールゲートリフター　　→　当該機器の操作実務経験、荷役作業の教育歴", None),
    ("　共通して強い資格　　　　→　労働安全衛生コンサルタント（安全／衛生）、RST講座 修了者", None),
    ("", None),
    ("■ この台帳の使い方", NAVY),
    ("・黄色のセルにご記入ください。科目名はシート02から自動で入ります。", None),
    ("・「修了証写しの確認」は、現物の写しを実際に確認した日を記入します。「持っているはず」では監督署に説明できません。", "B3402F"),
    ("・「台本レビュー完了日」は、監修者が実際に台本を読んで確認した日です。名前だけを載せる名義貸しは、講師要件を満たしたことになりません。", "B3402F"),
    ("・監修者の署名（または押印）を得て、教材ごとに保管してください。教材内および修了証にも監修者名を明示します。", None),
    ("", None),
    ("■ 根拠", NAVY),
    ("・昭和48年3月19日 基発第145号：「特別教育の講師については法令で資格要件は定められていないが、", None),
    ("　教習科目について十分な知識、経験を有する者でなければならないことは当然である」", None),
    ("・特別教育には登録・認可の制度がありません。事前に届け出る窓口はなく、労働基準監督署の臨検で事後的に確認されます。", None),
    ("・そのため、この台帳と台本のレビュー記録が、実務上ほぼ唯一の説明手段になります。", None),
]
for txt, col in GUIDE:
    cell(ws, r, 1, txt, border=False, size=9.5,
         bold=txt.startswith("■"), color=col or "4A5A68")
    r += 1
ws.freeze_panes = "A6"

wb.save("/home/user/-/estimate/聖建様_特別教育eラーニング_概算見積.xlsx")
print("saved")
